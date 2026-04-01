"""Gemini-based activity extraction from uploaded documents.

Supports PDF (sent as base64 inline data), Excel (converted to text),
and plain text / CSV files.
"""
from __future__ import annotations

import base64
import json
import logging
import re
import uuid
from pathlib import Path

import httpx

from app.models import Document, ExtractedActivity, Project

logger = logging.getLogger(__name__)

GEMINI_MODEL = "gemini-2.5-flash-preview-04-17"
GEMINI_URL = (
    "https://generativelanguage.googleapis.com/v1beta/models/"
    f"{GEMINI_MODEL}:generateContent"
)

VALID_UNIT_TYPES = {
    "Area", "AreaOverTime", "ContainerOverDistance", "Data", "DataOverTime",
    "Distance", "DistanceOverTime", "Energy", "Power", "Money", "Number",
    "NumberOverTime", "PassengerOverDistance", "Time", "Volume", "Weight",
    "WeightOverDistance", "WeightOverTime",
}

BRAND_MAP: dict[str, str] = {
    "nvidia": "computer servers",
    "gb200": "computer servers",
    "nvl72": "computer servers",
    "h100": "gpu servers",
    "a100": "gpu servers",
    "vertiv": "cooling equipment",
    "liebert": "cooling equipment",
    "caterpillar": "diesel generator",
    "cummins": "diesel generator",
    "tesla": "battery lithium ion",
    "megapack": "battery lithium ion",
    "powerpack": "battery lithium ion",
    "abb": "electrical switchgear",
    "schneider": "electrical equipment",
    "knauf": "insulation glass wool",
    "earthwool": "insulation glass wool",
    "kingspan": "insulation board",
}

EXTRACTION_PROMPT = """\
You are a carbon emissions data extraction specialist. Extract ALL emission-producing \
activities from the documents provided, formatted for the Climatiq emissions database.

FUNDAMENTAL RULE: Every activity missing quantity, unit, or unit_type will produce \
ZERO CO2e in the final carbon report. Use null ONLY as a last resort when the value \
is genuinely absent from the document — NEVER null just because you are uncertain. \
Estimate from context when needed.

=== STEP 1: SCAN FOR ALL QUANTITIES ===
Before writing any JSON, mentally locate EVERY quantity in the document:
- Tables: every numeric cell (kg, t, kWh, m2, units, $, count columns)
- Prose: "500 metric tonnes of...", "consuming 1,200 MWh annually"
- Counts: "32 x NVIDIA H100 servers" → quantity=32, unit="units", unit_type="Number"
- Power ratings: "2 x 500 kW generators" → quantity=1000, unit="kW", unit_type="Power"
- Cost/spend: "$3.2M capex" → quantity=3200000, unit="usd", unit_type="Money"
- Ranges: "200-250 t" → use lower bound: quantity=200
- Implied/annual values: look for associated numbers even if not directly adjacent

=== STEP 2: search_query (2-5 generic keywords, NO brand names) ===
The Climatiq Search API uses fuzzy matching. Verbose or brand-specific queries fail.
REQUIRED TRANSFORMATIONS:
- "NVIDIA H100 / GB200 / A100" → "gpu servers"
- "Vertiv Liebert XDU / cooling unit" → "cooling equipment"
- "Caterpillar C175 / Cummins QSK genset" → "diesel generator"
- "Tesla Megapack / Powerwall / PowerPack" → "battery lithium ion"
- "ABB switchgear / Schneider MV board" → "electrical switchgear"
- "Knauf / Earthwool / Kingspan insulation" → "insulation glass wool"
- "Portland cement CEM I/II/III" → "cement"
- "C30/C40 ready-mix / structural concrete" → "concrete"
- "Grade 60 rebar / BS4449 / deformed bar" → "steel rebar"
- "6061-T6 / 6063-T5 / extruded aluminum" → "aluminum sheet"
- "Grid electricity / mains power / utility" → "electricity supply grid"
- "Natural gas boiler / HVAC gas / gas-fired" → "natural gas combustion"
- "Diesel generator / diesel backup / diesel fuel use" → "diesel fuel combustion"
- "Road haulage / trucking / HGV / lorry" → "freight road truck"
- "Container shipping / sea freight / maritime" → "freight sea shipping"
- "Air cargo / air freight / airfreight" → "freight air transport"
Strip: all brand names, project codes (PRJ-123), phase/stage labels, \
adjectives (sustainable, low-carbon, green, premium, advanced, high-performance)

=== STEP 3: unit_type (EXACT case-sensitive value or null) ===
Weight     - kg, t, ton, tonne, lb, g, MT (mass of materials)
Energy     - kWh, MWh, GWh, MJ, GJ, TJ, therm (energy consumed)
Power      - W, kW, MW, GW (installed/rated capacity, NOT energy over time)
Volume     - l, L, m3, gallon, bbl, ft3 (fluid/gas volumes)
Area       - m2, km2, ft2, acres (surface area)
Distance   - km, mi, miles (travel or transport distances)
Money      - $, USD, EUR, GBP, AUD and other currencies (spend amounts)
Number     - units, pieces, items, servers, vehicles, count (discrete items)
Data       - MB, GB, TB, PB (digital storage/transfer)
Time       - hours, days, months, years (operational duration)
WeightOverDistance   - t*km, tonne-km (freight transport)
PassengerOverDistance - passenger-km, pax*km
AreaOverTime         - m2/year, m2*year

=== STEP 4: DUAL ROW RULE ===
If a line has BOTH a physical quantity AND a monetary value, create TWO rows:
  Row 1 (physical): unit_type=<physical>, quantity=<physical amount>, unit=<physical unit>
  Row 2 (monetary): unit_type="Money", quantity=<spend>, unit=<currency code e.g. "usd">

=== STEP 5: CATEGORY ===
HARDWARE | CONSTRUCTION | ENERGY | TRANSPORT | OPERATIONS | PROCUREMENT | WASTE | WATER | OTHER

=== COMPLETE EXAMPLE OUTPUT ===
[
  {
    "search_query": "cement",
    "raw_text": "Portland Cement CEM I 42.5N - 850 metric tonnes",
    "unit_type": "Weight",
    "quantity": 850,
    "unit": "t",
    "amount": null,
    "currency": null,
    "region": null,
    "category": "CONSTRUCTION",
    "confidence": "HIGH",
    "source_doc_index": 0,
    "source_page": "BOM row 14",
    "note": null
  },
  {
    "search_query": "electricity supply grid",
    "raw_text": "Annual electricity consumption 4,800 MWh @ $0.12/kWh = $576,000",
    "unit_type": "Energy",
    "quantity": 4800,
    "unit": "MWh",
    "amount": null,
    "currency": null,
    "region": null,
    "category": "ENERGY",
    "confidence": "HIGH",
    "source_doc_index": 0,
    "source_page": "p.12 Table 3",
    "note": null
  },
  {
    "search_query": "gpu servers",
    "raw_text": "96 x NVIDIA H100 80GB SXM5, capex $18.5M",
    "unit_type": "Number",
    "quantity": 96,
    "unit": "units",
    "amount": null,
    "currency": null,
    "region": null,
    "category": "HARDWARE",
    "confidence": "HIGH",
    "source_doc_index": 0,
    "source_page": "p.5",
    "note": "Monetary fallback row generated separately"
  },
  {
    "search_query": "gpu servers",
    "raw_text": "96 x NVIDIA H100 80GB SXM5, capex $18.5M",
    "unit_type": "Money",
    "quantity": 18500000,
    "unit": "usd",
    "amount": 18500000,
    "currency": "usd",
    "region": null,
    "category": "HARDWARE",
    "confidence": "MEDIUM",
    "source_doc_index": 0,
    "source_page": "p.5",
    "note": "Monetary fallback for H100 server capex"
  },
  {
    "search_query": "diesel generator",
    "raw_text": "2 x 500 kW Cummins backup diesel generators",
    "unit_type": "Power",
    "quantity": 1000,
    "unit": "kW",
    "amount": null,
    "currency": null,
    "region": null,
    "category": "HARDWARE",
    "confidence": "HIGH",
    "source_doc_index": 0,
    "source_page": "spec sheet p.3",
    "note": "Two 500 kW units combined"
  }
]

Now extract ALL emission-producing activities from the provided documents and return \
a valid JSON array in EXACTLY the format above. No markdown fences, no explanation."""


def _validate_unit_type(ut: str | None) -> str | None:
    if not ut:
        return None
    if ut in VALID_UNIT_TYPES:
        return ut
    for valid in VALID_UNIT_TYPES:
        if valid.lower() == ut.lower():
            return valid
    return None


def _normalize_search_query(query: str) -> str:
    if not query:
        return "unknown activity"
    q = query.lower().strip()
    q = re.sub(r"\b(phase|stage|q[1-4]|fy\d+|20\d{2})\s*\d*", "", q, flags=re.IGNORECASE)
    q = re.sub(
        r"\b(sustainable|low-carbon|eco-friendly|green|high-performance|premium|advanced)\b",
        "",
        q,
        flags=re.IGNORECASE,
    )
    q = re.sub(
        r"\b(procurement|supply|provision|installation|of|the|and|for|from|with)\b",
        " ",
        q,
        flags=re.IGNORECASE,
    )
    words = q.split()
    for word in words:
        if word in BRAND_MAP:
            return BRAND_MAP[word]
    cleaned = " ".join(words[:5])
    return cleaned or "unknown activity"


def _excel_to_text(path: Path) -> str:
    """Convert an Excel workbook to a plain-text table for Gemini."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sections: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        sections.append(f"Sheet: {sheet_name}")
        for row in rows:
            sections.append("\t".join("" if c is None else str(c) for c in row))
    wb.close()
    return "\n".join(sections)


def _build_parts(
    docs: list[Document],
    stored_paths: list[Path | None],
) -> list[dict]:
    """Build the Gemini API `parts` list from the uploaded documents."""
    parts: list[dict] = []

    for idx, (doc, path) in enumerate(zip(docs, stored_paths)):
        label = f"[Document {idx}: {doc.filename}]"

        if path is None or not path.exists():
            parts.append({"text": f"{label}\n[File not found on disk]"})
            continue

        ext = path.suffix.lower()

        if ext == ".pdf":
            raw_bytes = path.read_bytes()
            parts.append({
                "inlineData": {
                    "mimeType": "application/pdf",
                    "data": base64.b64encode(raw_bytes).decode("utf-8"),
                }
            })
            # Label comes after so Gemini sees the index
            parts.append({"text": label})

        elif ext in (".xlsx", ".xls"):
            try:
                text_repr = _excel_to_text(path)
            except Exception as exc:
                text_repr = f"[Excel parse error: {exc}]"
            parts.append({"text": f"{label}\n{text_repr[:30000]}"})

        else:
            # Plain text, CSV, TSV, etc.
            try:
                content = path.read_text(encoding="utf-8", errors="replace")
            except Exception as exc:
                content = f"[Read error: {exc}]"
            parts.append({"text": f"{label}\n{content[:20000]}"})

    return parts


_INFER_QUANTITIES_PROMPT = """\
Some emission activities were extracted from the documents above but are missing \
a quantity and/or unit_type. Using the document context, fill in the missing values.

Activities needing quantity/unit inference (JSON):
{items_json}

For each activity:
1. Search the documents for any number directly associated with this activity.
   If found: use the exact value (confidence="high").
2. If no number is present: use a typical/representative amount for this type of activity
   (confidence="low").
   Examples: electricity → 1000 kWh, cement → 1 t, diesel → 1000 L, flight → 5000 km,
   server → 1 units, natural gas → 10000 m3, road freight → 10000 t*km

unit_type must be exactly one of:
Weight, Energy, Power, Volume, Area, Distance, Money, Number, Data, Time,
WeightOverDistance, PassengerOverDistance, AreaOverTime, DataOverTime,
DistanceOverTime, NumberOverTime, WeightOverTime

Respond with ONLY a JSON array (no markdown fences), preserving the input order:
[{{"id": "act_0001", "quantity": 100.0, "unit": "t", "unit_type": "Weight", \
"confidence": "high|medium|low", "rationale": "one sentence"}}]"""


async def _infer_missing_quantities(
    client: httpx.AsyncClient,
    api_key: str,
    activities: list[ExtractedActivity],
    document_parts: list[dict],
) -> list[ExtractedActivity]:
    """Second-pass Gemini call to fill in quantity/unit_type for activities where both are null."""
    missing = [
        a for a in activities
        if a.quantity is None and a.amount is None and a.unit_type is None
    ]
    if not missing:
        return activities

    items = [
        {"id": a.id, "raw_text": a.note or a.text, "search_query": a.text}
        for a in missing
    ]
    prompt = _INFER_QUANTITIES_PROMPT.format(items_json=json.dumps(items, ensure_ascii=False))
    all_parts = list(document_parts) + [{"text": prompt}]

    try:
        resp = await client.post(
            GEMINI_URL,
            params={"key": api_key},
            json={
                "contents": [{"parts": all_parts}],
                "generationConfig": {
                    "temperature": 0.1,
                    "maxOutputTokens": 4096,
                    "response_mime_type": "application/json",
                },
            },
        )
        resp.raise_for_status()
        data = resp.json()
    except Exception as exc:
        logger.warning("Gemini quantity inference call failed: %s", exc)
        return activities

    raw_text: str = (
        data.get("candidates", [{}])[0]
        .get("content", {})
        .get("parts", [{}])[0]
        .get("text", "")
    )
    raw_text = re.sub(r"```json\s*", "", raw_text)
    raw_text = re.sub(r"```\s*", "", raw_text).strip()

    try:
        inferred: list[dict] = json.loads(raw_text)
        if isinstance(inferred, dict):
            for key in ("activities", "items", "results", "data"):
                if key in inferred and isinstance(inferred[key], list):
                    inferred = inferred[key]
                    break
    except Exception as exc:
        logger.warning("Gemini quantity inference parse failed: %s", exc)
        return activities

    inferred_by_id: dict[str, dict] = {
        item.get("id"): item for item in inferred if isinstance(item, dict) and item.get("id")
    }

    updated: list[ExtractedActivity] = []
    for activity in activities:
        inf = inferred_by_id.get(activity.id)
        if inf and activity.quantity is None and activity.amount is None and activity.unit_type is None:
            qty = _to_float(inf.get("quantity"))
            ut = _validate_unit_type(inf.get("unit_type"))
            unit = inf.get("unit") or None
            if qty is not None or ut is not None:
                logger.info(
                    "Inferred missing quantity for '%s': %s %s (%s)",
                    activity.text[:60], qty, unit or ut, inf.get("confidence", "?"),
                )
                updated.append(ExtractedActivity(
                    id=activity.id,
                    projectId=activity.projectId,
                    text=activity.text,
                    search_query=activity.search_query,
                    unit_type=ut,
                    region=activity.region,
                    quantity=qty,
                    unit=unit,
                    amount=activity.amount,
                    currency=activity.currency,
                    sourceDocumentId=activity.sourceDocumentId,
                    note=activity.note,
                ))
                continue
        updated.append(activity)

    return updated


async def extract_with_gemini(
    api_key: str,
    project: Project,
    docs: list[Document],
    stored_paths: list[Path | None],
) -> list[ExtractedActivity]:
    """Call Gemini to extract activities from the provided documents.

    Falls back to an empty list if Gemini returns an unusable response.
    """
    document_parts = _build_parts(docs, stored_paths)
    if not document_parts:
        return []

    # Instruction comes last so Gemini reads the documents first
    all_parts = document_parts + [{"text": EXTRACTION_PROMPT}]

    async with httpx.AsyncClient(timeout=180.0) as client:
        resp = await client.post(
            GEMINI_URL,
            params={"key": api_key},
            json={
                "contents": [{"parts": all_parts}],
                "generationConfig": {
                    "temperature": 0.1,
                    "maxOutputTokens": 16384,
                    "response_mime_type": "application/json",
                },
            },
        )
        resp.raise_for_status()
        data = resp.json()

        raw_text: str = (
            data.get("candidates", [{}])[0]
            .get("content", {})
            .get("parts", [{}])[0]
            .get("text", "")
        )
        raw_text = re.sub(r"```json\s*", "", raw_text)
        raw_text = re.sub(r"```\s*", "", raw_text).strip()

        extracted: list[dict] = []
        if raw_text:
            try:
                parsed = json.loads(raw_text)
                if isinstance(parsed, list):
                    extracted = parsed
                elif isinstance(parsed, dict):
                    # JSON-mode may wrap in an object — check common keys
                    for key in ("activities", "items", "results", "data", "rows"):
                        if key in parsed and isinstance(parsed[key], list):
                            extracted = parsed[key]
                            break
            except Exception:
                pass

        activities: list[ExtractedActivity] = []
        for idx, item in enumerate(extracted):
            src_idx = item.get("source_doc_index") or 0
            src_doc = docs[src_idx] if isinstance(src_idx, int) and src_idx < len(docs) else None

            raw = item.get("raw_text") or ""
            sq = _normalize_search_query(item.get("search_query") or raw)

            raw_cat = item.get("category") or None
            category = raw_cat.upper() if raw_cat else None

            activities.append(
                ExtractedActivity(
                    id=str(uuid.uuid4()),
                    projectId=project.id,
                    text=sq or "Unknown activity",
                    search_query=sq or None,
                    unit_type=_validate_unit_type(item.get("unit_type")),
                    region=item.get("region") or project.primaryRegion,
                    quantity=_to_float(item.get("quantity")),
                    unit=item.get("unit") or None,
                    amount=_to_float(item.get("amount")),
                    currency=item.get("currency") or None,
                    sourceDocumentId=src_doc.id if src_doc else None,
                    note=(raw[:200] if raw else item.get("note") or None),
                    category=category,
                )
            )

        # Second pass: infer missing quantity/unit_type for activities that slipped through
        activities = await _infer_missing_quantities(client, api_key, activities, document_parts)

    return activities


def _to_float(val: object) -> float | None:
    if val is None:
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


PHASE_CLASSIFICATION_PROMPT = """\
You are a carbon accounting expert specialising in AI data-center infrastructure,
following the GHG Protocol embodied/operational distinction.

DEFINITIONS (apply strictly in this order):
1. EMBODIED — one-time capital expenditures for setting up AI-enabling infrastructure:
   • Server / GPU / accelerator procurement (H100, A100, GB200, DGX, etc.)
   • Cooling equipment installation (CRAC/CRAH units, liquid cooling manifolds, CDUs)
   • Electrical & mechanical infrastructure (PDUs, switchgear, UPS, transformers, generators)
   • Data-center construction, fit-out, civil works, steel, concrete, cabling
   • Network hardware (switches, routers, optics — one-time)
   • Battery / energy-storage systems (Megapack, Powerpack)
   • Any item described with: purchased, procured, installed, deployed, manufactured,
     acquired, capex, one-time, units, servers, racks, modules

2. OPERATIONAL — ongoing/recurring costs of running the data center:
   • Electricity consumption (kWh, MWh, GWh — grid or on-site solar)
   • Cooling energy use (chiller runtime, HVAC, mechanical ventilation)
   • Diesel or natural-gas for backup generators / heating (fuel volumes or energy)
   • Water consumption for cooling towers
   • Network bandwidth / data transfer charges
   • Cloud or co-location service fees (recurring SaaS, IaaS)
   • Maintenance, support contracts, cleaning, staffing
   • Any item described with: annual, monthly, per year, consumption, usage, billing,
     ongoing, recurring, opex, kWh, MWh

KEY RULE: If a quantity has physical units of kWh / MWh / GWh / litre / m³ / gallon
(i.e. metered consumption) → ALWAYS operational.
If a quantity is a count of physical assets (units, servers, racks) AND the description
is about procurement/purchase → ALWAYS embodied.

Return a JSON array, one object per input activity (include ALL ids):
[{"id": "<activity_id>", "phase": "embodied" | "operational", "reason": "<one sentence>"}]

Activities to classify:
"""


async def classify_phases_with_gemini(
    api_key: str,
    activities: list[ExtractedActivity],
) -> dict[str, dict]:
    """Call Gemini to classify each activity as embodied or operational.

    Returns a dict mapping activity_id -> {"phase": str, "reason": str}.
    """
    if not activities:
        return {}

    activity_list = [
        {
            "id": act.id,
            "text": act.text,
            "category": act.category,
            "quantity": act.quantity,
            "unit": act.unit,
        }
        for act in activities
    ]

    prompt = PHASE_CLASSIFICATION_PROMPT + json.dumps(activity_list, indent=2)

    async with httpx.AsyncClient(timeout=60.0) as client:
        resp = await client.post(
            GEMINI_URL,
            params={"key": api_key},
            json={
                "contents": [{"parts": [{"text": prompt}]}],
                "generationConfig": {
                    "temperature": 0.1,
                    "maxOutputTokens": 8192,
                    "response_mime_type": "application/json",
                },
            },
        )
        resp.raise_for_status()
        data = resp.json()

    raw_text: str = (
        data.get("candidates", [{}])[0]
        .get("content", {})
        .get("parts", [{}])[0]
        .get("text", "")
    )
    raw_text = re.sub(r"```json\s*", "", raw_text)
    raw_text = re.sub(r"```\s*", "", raw_text).strip()

    result: dict[str, dict] = {}
    if raw_text:
        try:
            parsed = json.loads(raw_text)
            if isinstance(parsed, list):
                for item in parsed:
                    act_id = item.get("id")
                    phase = item.get("phase")
                    reason = item.get("reason", "")
                    if act_id and phase in ("embodied", "operational"):
                        result[act_id] = {"phase": phase, "reason": reason}
        except Exception:
            pass

    return result
